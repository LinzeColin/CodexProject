#!/usr/bin/env python3
"""KMFA v1.5 S06-P1 authoritative source registration.

This phase is explicitly allowed to read/list/stat/hash and parse the one
private authority package required by the v1.5 Roadmap.  Raw names, hashes,
text, displayed values, formulas, sheet names, and locators are written only
to a mode-0600 file below the Git-ignored private runtime.  Tracked callers
consume only :func:`public_projection`.

The phase registers candidates; it does not approve a golden value.  OCR,
formula results, cached display values, and manually entered values remain
distinct and cannot become final facts in S06-P1.
"""

from __future__ import annotations

import argparse
import hashlib
import json
import os
import re
import stat
import unicodedata
import warnings
import zipfile
from collections import Counter
from datetime import date, datetime
from io import BytesIO
from pathlib import Path
from typing import Any, Iterable


RUN_PHASE_ID = "V015_S06_P1_AUTHORITATIVE_SOURCE_REGISTRATION"
ROADMAP_PHASE_ID = "S06-P1"
TASK_ID = "KMFA-V015-S06-P1-AUTHORITATIVE-SOURCE-REGISTRATION-20260715"
ACCEPTANCE_ID = "ACC-KMFA-V015-S06-P1-AUTHORITATIVE-SOURCE-REGISTRATION"
VERSION = "1.5.0-dev-s06p1"
PRIVATE_SCHEMA_VERSION = "kmfa.private.v015.s06p1.authoritative_source_registration.v1"
PUBLIC_SCHEMA_VERSION = "kmfa.v015.s06p1.authoritative_source_registration_public_safe.v1"

RAW_INBOX = Path.home() / "Downloads" / "KMFA_MetaData"
PRIVATE_PACKAGE_ENV = "KMFA_V015_S06_P1_AUTHORITY_PACKAGE"
PRIVATE_OUTPUT_DIR = Path("KMFA/.codex_private_runtime/v015_s06_p1_authoritative_source_registration")
PRIVATE_MANIFEST_PATH = PRIVATE_OUTPUT_DIR / "private_authority_registration.json"

EXPECTED_SOURCE_COUNT = 9
EXPECTED_PDF_COUNT = 8
EXPECTED_WORKBOOK_COUNT = 1
FIELD_FAMILIES = (
    "PROJECT_IDENTITY",
    "CONTRACT_AMOUNT",
    "TOTAL_EXPENDITURE",
    "GROSS_PROFIT",
    "GROSS_MARGIN",
    "COST_CATEGORY",
)

# These are public requirement vocabulary and generic template aliases.  They
# do not contain any private source-specific header, project, or customer text.
FIELD_ALIASES: dict[str, tuple[str, ...]] = {
    "PROJECT_IDENTITY": ("项目名称", "项目编码", "合同编号", "合同号", "公司名称", "甲方名称"),
    "CONTRACT_AMOUNT": ("合同额", "合同金额", "含税合同金额"),
    "TOTAL_EXPENDITURE": ("资金运用及各项支出", "总支出", "总成本", "合计支出"),
    "GROSS_PROFIT": ("毛利",),
    "GROSS_MARGIN": ("毛利率",),
    "COST_CATEGORY": (
        "原材料", "租赁费", "机械费", "保险费", "现场管理费", "工资", "税金", "信息费",
        "管理费用", "占用的资金利息", "生活费用", "工程车辆使用费", "差旅费", "业务费用",
        "招待费", "生活用品", "办公费", "水电费", "安全防护费",
    ),
}

# Workbook column binding is intentionally exact.  Substring matching made
# ``毛利率`` look like both gross profit and margin, and treated unrelated
# amount columns as contract amounts.  These semantic headers are public
# template vocabulary; private values remain in the ignored runtime.
WORKBOOK_HEADER_SEMANTICS: dict[str, tuple[str, str | None]] = {
    "公司名称": ("PROJECT_IDENTITY", "COUNTERPARTY_NAME"),
    "甲方名称": ("PROJECT_IDENTITY", "COUNTERPARTY_NAME"),
    "项目名称": ("PROJECT_IDENTITY", "PROJECT_NAME"),
    "项目编码": ("PROJECT_IDENTITY", "PROJECT_ID"),
    "合同编号": ("PROJECT_IDENTITY", "CONTRACT_ID"),
    "合同号": ("PROJECT_IDENTITY", "CONTRACT_ID"),
    "合同额": ("CONTRACT_AMOUNT", None),
    "合同金额": ("CONTRACT_AMOUNT", None),
    "含税合同金额": ("CONTRACT_AMOUNT", None),
    "总支出": ("TOTAL_EXPENDITURE", None),
    "总成本": ("TOTAL_EXPENDITURE", None),
    "合计支出": ("TOTAL_EXPENDITURE", None),
    "毛利": ("GROSS_PROFIT", None),
    "毛利润": ("GROSS_PROFIT", None),
    "毛利率": ("GROSS_MARGIN", None),
}

_PDF_TOP_LEVEL_CATEGORY = re.compile(
    r"^(?:[（(][一二三四五六七八九十]+[）)]|三?1[.。][12])"
)
_WORKBOOK_SUMMARY_LABELS = frozenset({"总和", "合计"})
_CANDIDATE_ROLES = frozenset({
    "IDENTITY_COMPONENT", "PRIMARY_FIELD", "CROSS_SOURCE_PRIMARY_FIELD",
    "INTERMEDIATE_TOTAL", "TOP_LEVEL_CATEGORY",
})

TEMPLATE_STRATEGIES: dict[str, str] = {
    "PDF_CURRENT_COMPACT": "TEXT_LAYER_PRIMARY_TABLE_WITH_COMPACT_ROW_STRATEGY",
    "PDF_CURRENT_COMPACT_WITH_SUPPORTING_APPENDIX": "PRIMARY_TABLE_PLUS_PAGE_SCOPED_SUPPORTING_APPENDIX_STRATEGY",
    "PDF_LEGACY_DETAILED": "TEXT_LAYER_DETAILED_COST_ROW_STRATEGY",
    "XLSX_FORMULA_SUMMARY": "FORMULA_AND_CACHED_DISPLAY_DUAL_READ_STRATEGY",
    "XLSX_MANUAL_DETAIL": "TYPED_CELL_AND_HEADER_LOCATOR_STRATEGY",
    "XLSX_IMAGE_EMBEDDED": "IMAGE_COMPONENT_QUARANTINE_AND_HUMAN_REVIEW_STRATEGY",
}

_NUMBER_TOKEN = re.compile(r"(?<![A-Za-z0-9])[-+]?\d[\d,， ]*(?:\.\d+)?%?")


class RegistrationError(RuntimeError):
    """Fail-closed S06-P1 registration error."""


def _sha256_bytes(payload: bytes) -> str:
    return hashlib.sha256(payload).hexdigest()


def _sha256_text(value: str) -> str:
    return hashlib.sha256(value.encode("utf-8")).hexdigest()


def _sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _normalize(value: Any) -> str:
    return re.sub(r"\s+", "", unicodedata.normalize("NFKC", str(value or ""))).casefold()


def _json_value(value: Any) -> Any:
    if value is None or isinstance(value, (bool, int, float, str)):
        return value
    if isinstance(value, (date, datetime)):
        return value.isoformat()
    return str(value)


def _stat_snapshot(path: Path) -> dict[str, int]:
    value = path.stat()
    return {
        "device": value.st_dev,
        "inode": value.st_ino,
        "mode": stat.S_IMODE(value.st_mode),
        "size_bytes": value.st_size,
        "mtime_ns": value.st_mtime_ns,
        "ctime_ns": value.st_ctime_ns,
    }


def _hidden_member(name: str) -> bool:
    return "__MACOSX/" in name or Path(name).name.startswith(".")


def _business_members(archive: zipfile.ZipFile) -> list[zipfile.ZipInfo]:
    return [
        info for info in archive.infolist()
        if not info.is_dir() and not _hidden_member(info.filename)
    ]


def inspect_archive_shape(path: Path) -> dict[str, int] | None:
    try:
        with zipfile.ZipFile(path) as archive:
            members = _business_members(archive)
            return {
                "source_count": len(members),
                "pdf_count": sum(Path(item.filename).suffix.lower() == ".pdf" for item in members),
                "workbook_count": sum(Path(item.filename).suffix.lower() in {".xlsx", ".xlsm"} for item in members),
                "hidden_member_count": sum(
                    not item.is_dir() and _hidden_member(item.filename) for item in archive.infolist()
                ),
            }
    except (OSError, zipfile.BadZipFile):
        return None


def resolve_authority_package(raw_root: Path = RAW_INBOX) -> tuple[Path | None, str, int]:
    env_value = os.environ.get(PRIVATE_PACKAGE_ENV, "").strip()
    if env_value:
        candidate = Path(env_value)
        if not candidate.is_absolute():
            candidate = raw_root / candidate
        return (
            candidate if candidate.is_file() else None,
            "ENV_PATH_FOUND" if candidate.is_file() else "ENV_PATH_MISSING",
            1,
        )
    if not raw_root.is_dir():
        return None, "RAW_ROOT_UNAVAILABLE", 0
    matches: list[Path] = []
    for candidate in sorted(raw_root.rglob("*.zip")):
        shape = inspect_archive_shape(candidate)
        if shape and (
            shape["source_count"], shape["pdf_count"], shape["workbook_count"]
        ) == (EXPECTED_SOURCE_COUNT, EXPECTED_PDF_COUNT, EXPECTED_WORKBOOK_COUNT):
            matches.append(candidate)
    if len(matches) == 1:
        return matches[0], "PUBLIC_SHAPE_UNIQUE_MATCH", 1
    if matches:
        return None, "PUBLIC_SHAPE_AMBIGUOUS", len(matches)
    return None, "PUBLIC_SHAPE_NO_MATCH", 0


def _families_for_text(text: str) -> set[str]:
    normalized = _normalize(text)
    families: set[str] = set()
    for family, aliases in FIELD_ALIASES.items():
        if any(_normalize(alias) in normalized for alias in aliases):
            families.add(family)
    # ``毛利率`` contains the substring ``毛利`` but is not a gross-profit
    # amount.  A row explicitly labelled 毛利 and containing a percentage still
    # yields both amount and margin candidates below.
    if _normalize("毛利率") in normalized and _normalize("毛利") not in normalized.replace(
        _normalize("毛利率"), "",
    ):
        families.discard("GROSS_PROFIT")
        families.add("GROSS_MARGIN")
    # A cost-table gross-profit row normally contains both the amount and its
    # ratio.  It yields two candidates from one raw line; S06-P2 must confirm
    # which numeric token belongs to which field.
    if "GROSS_PROFIT" in families and any(token.endswith("%") for token in _NUMBER_TOKEN.findall(text)):
        families.add("GROSS_MARGIN")
    return families


def _pdf_families_for_line(text: str) -> set[str]:
    """Return primary-table candidate families for one PDF text-layer line.

    PDF tables contain nested detail lines and supporting appendices.  The
    golden baseline needs the primary project table, so this classifier keeps
    explicit identity/amount/total/profit rows plus top-level cost categories.
    It never assigns contract and total expenditure to the same line.
    """

    normalized = _normalize(text)
    families: set[str] = set()
    if any(_normalize(alias) in normalized for alias in FIELD_ALIASES["PROJECT_IDENTITY"]):
        families.add("PROJECT_IDENTITY")
    if any(_normalize(alias) in normalized for alias in FIELD_ALIASES["CONTRACT_AMOUNT"]):
        families.add("CONTRACT_AMOUNT")
    elif any(_normalize(alias) in normalized for alias in FIELD_ALIASES["TOTAL_EXPENDITURE"]):
        families.add("TOTAL_EXPENDITURE")
    if _normalize("毛利") in normalized:
        if _normalize("毛利率") in normalized and _normalize("毛利") not in normalized.replace(
            _normalize("毛利率"), "",
        ):
            families.add("GROSS_MARGIN")
        else:
            families.add("GROSS_PROFIT")
            if any(token.endswith("%") for token in _NUMBER_TOKEN.findall(text)):
                families.add("GROSS_MARGIN")
    if _PDF_TOP_LEVEL_CATEGORY.match(normalized) and any(
        _normalize(alias) in normalized for alias in FIELD_ALIASES["COST_CATEGORY"]
    ):
        families.add("COST_CATEGORY")
    return families


def _candidate_role(family: str, text: str, *, workbook: bool = False) -> str:
    normalized = _normalize(text)
    if family == "PROJECT_IDENTITY":
        return "IDENTITY_COMPONENT"
    if family == "COST_CATEGORY":
        return "TOP_LEVEL_CATEGORY"
    if family == "TOTAL_EXPENDITURE" and _normalize("合计支出") not in normalized:
        return "INTERMEDIATE_TOTAL"
    if workbook:
        return "CROSS_SOURCE_PRIMARY_FIELD"
    return "PRIMARY_FIELD"


def _identity_component(text: str) -> str | None:
    normalized = _normalize(text)
    for label, component in (
        ("合同编号", "CONTRACT_ID"), ("合同号", "CONTRACT_ID"),
        ("项目编码", "PROJECT_ID"), ("项目名称", "PROJECT_NAME"),
        ("公司名称", "COUNTERPARTY_NAME"), ("甲方名称", "COUNTERPARTY_NAME"),
    ):
        if _normalize(label) in normalized:
            return component
    return None


def _private_text_candidate(
    *, source_ref: str, locator: str, family: str, raw_text: str, method: str,
    formula_text: Any = None, cached_display_value: Any = None,
    candidate_role: str | None = None, identity_component: str | None = None,
) -> dict[str, Any]:
    return {
        "source_ref": source_ref,
        "source_locator": locator,
        "field_family": family,
        "raw_text": raw_text,
        "original_display_tokens": _NUMBER_TOKEN.findall(raw_text),
        "formula_text": _json_value(formula_text),
        "cached_display_value": _json_value(cached_display_value),
        "extraction_method": method,
        "candidate_role": candidate_role or _candidate_role(family, raw_text),
        "identity_component": identity_component,
        "candidate_status": "CANDIDATE_NOT_FINAL",
        "human_confirmation_required": True,
        "ocr_final_fact_allowed": False,
    }


def inspect_pdf_payload(payload: bytes, source_ref: str) -> dict[str, Any]:
    try:
        from pypdf import PdfReader  # type: ignore
    except ImportError as error:  # pragma: no cover - environment-specific gate
        raise RegistrationError("pypdf is required for the private PDF scan") from error

    try:
        reader = PdfReader(BytesIO(payload), strict=False)
    except Exception as error:
        raise RegistrationError(f"private PDF is not readable for {source_ref}: {type(error).__name__}") from error

    pages: list[dict[str, Any]] = []
    candidates: list[dict[str, Any]] = []
    quarantined: list[dict[str, Any]] = []
    for page_index, page in enumerate(reader.pages, start=1):
        try:
            text = page.extract_text() or ""
        except Exception as error:  # fail this component closed without losing the other sources
            text = ""
            quarantined.append({"locator": f"PAGE_{page_index}", "reason": f"TEXT_EXTRACTION_{type(error).__name__}"})
        lines = [line.strip() for line in text.splitlines() if line.strip()]
        try:
            image_count = len(page.images)
        except Exception:
            image_count = 0
        page_families: Counter[str] = Counter()
        for line_index, line in enumerate(lines, start=1):
            # Supporting pages are retained verbatim in ``pages`` but are not
            # promoted into the primary-table golden candidate set.
            if page_index != 1:
                continue
            for family in sorted(_pdf_families_for_line(line)):
                page_families[family] += 1
                candidates.append(_private_text_candidate(
                    source_ref=source_ref,
                    locator=f"PAGE_{page_index}:LINE_{line_index}",
                    family=family,
                    raw_text=line,
                    method="PDF_TEXT_LAYER",
                    candidate_role=_candidate_role(family, line),
                    identity_component=_identity_component(line) if family == "PROJECT_IDENTITY" else None,
                ))
        if not text.strip():
            quarantined.append({"locator": f"PAGE_{page_index}", "reason": "TEXTLESS_PAGE_REQUIRES_HUMAN_REVIEW"})
        pages.append({
            "page_index": page_index,
            "raw_text": text,
            "text_character_count": len(text),
            "nonempty_line_count": len(lines),
            "embedded_image_count": image_count,
            "field_family_hit_counts": dict(sorted(page_families.items())),
            "text_layer_present": bool(text.strip()),
        })

    primary_lines = pages[0]["nonempty_line_count"] if pages else 0
    primary_family_count = sum(pages[0]["field_family_hit_counts"].values()) if pages else 0
    if not pages or not pages[0]["text_layer_present"] or primary_family_count < 3:
        template_class = "UNRECOGNIZED_PDF_TEMPLATE"
    elif primary_lines <= 36:
        template_class = "PDF_CURRENT_COMPACT_WITH_SUPPORTING_APPENDIX" if len(pages) > 1 else "PDF_CURRENT_COMPACT"
    else:
        template_class = "PDF_LEGACY_DETAILED"
    if template_class not in TEMPLATE_STRATEGIES:
        quarantined.append({"locator": "SOURCE", "reason": "UNRECOGNIZED_TEMPLATE"})
    return {
        "source_ref": source_ref,
        "format": "PDF",
        "page_count": len(pages),
        "text_layer_page_count": sum(item["text_layer_present"] for item in pages),
        "textless_page_count": sum(not item["text_layer_present"] for item in pages),
        "embedded_image_count": sum(item["embedded_image_count"] for item in pages),
        "pages": pages,
        "field_candidates": candidates,
        "template_class": template_class,
        "template_generation_status": "STRUCTURAL_CANDIDATE_REQUIRES_HUMAN_CONFIRMATION",
        "parser_strategy": TEMPLATE_STRATEGIES.get(template_class, "QUARANTINE_ONLY"),
        "quarantined_components": quarantined,
        "ocr_performed": False,
        "ocr_final_fact_count": 0,
    }


def _xlsx_media_summary(payload: bytes) -> dict[str, Any]:
    with zipfile.ZipFile(BytesIO(payload)) as archive:
        names = archive.namelist()
        media = [name for name in names if name.startswith("xl/media/") and not name.endswith("/")]
        drawings = [name for name in names if name.startswith("xl/drawings/") and not name.endswith("/")]
        return {
            "media_count": len(media),
            "media_suffixes": sorted({Path(name).suffix.lower() for name in media}),
            "drawing_part_count": len(drawings),
            "calculation_chain_present": "xl/calcChain.xml" in names,
        }


def inspect_workbook_payload(payload: bytes, source_ref: str) -> dict[str, Any]:
    try:
        from openpyxl import load_workbook  # type: ignore
    except ImportError as error:  # pragma: no cover - environment-specific gate
        raise RegistrationError("openpyxl is required for the private workbook scan") from error
    try:
        with warnings.catch_warnings():
            warnings.simplefilter("ignore")
            workbook = load_workbook(BytesIO(payload), read_only=False, data_only=False)
            displayed = load_workbook(BytesIO(payload), read_only=False, data_only=True)
    except Exception as error:
        raise RegistrationError(f"private workbook is not readable for {source_ref}: {type(error).__name__}") from error

    media = _xlsx_media_summary(payload)
    sheet_records: list[dict[str, Any]] = []
    candidates: list[dict[str, Any]] = []
    formula_count = cached_formula_count = manual_numeric_count = 0
    for sheet_index, (sheet, display_sheet) in enumerate(zip(workbook.worksheets, displayed.worksheets), start=1):
        nonempty_count = sheet_formula_count = sheet_cached_count = sheet_manual_numeric = 0
        header_bindings: list[tuple[int, int, str, str, str | None]] = []
        for row in sheet.iter_rows():
            for cell in row:
                value = cell.value
                if value is not None:
                    nonempty_count += 1
                if isinstance(value, str) and value.startswith("="):
                    sheet_formula_count += 1
                    if display_sheet[cell.coordinate].value is not None:
                        sheet_cached_count += 1
                elif isinstance(value, (int, float)) and not isinstance(value, bool):
                    sheet_manual_numeric += 1
                if isinstance(value, str) and not value.startswith("="):
                    semantic = WORKBOOK_HEADER_SEMANTICS.get(_normalize(value))
                    if semantic is not None:
                        family, component = semantic
                        header_bindings.append((cell.row, cell.column, family, value, component))

        for header_row, column, family, header_text, component in header_bindings:
            # A table ends at the first completely blank row.  This prevents a
            # header in one repeated block from consuming later headers,
            # totals, or unrelated sections in the same column.
            end_row = sheet.max_row
            for candidate_end in range(header_row + 1, sheet.max_row + 1):
                if all(sheet.cell(row=candidate_end, column=index).value is None for index in range(1, sheet.max_column + 1)):
                    end_row = candidate_end - 1
                    break
            for row_index in range(header_row + 1, end_row + 1):
                cell = sheet.cell(row=row_index, column=column)
                if cell.value is None:
                    continue
                row_labels = {
                    _normalize(sheet.cell(row=row_index, column=index).value)
                    for index in range(1, sheet.max_column + 1)
                    if sheet.cell(row=row_index, column=index).value is not None
                }
                if row_labels & _WORKBOOK_SUMMARY_LABELS:
                    continue
                normalized_value = _normalize(cell.value)
                if normalized_value in _WORKBOOK_SUMMARY_LABELS:
                    continue
                if WORKBOOK_HEADER_SEMANTICS.get(normalized_value) is not None:
                    continue
                display_value = display_sheet[cell.coordinate].value
                candidates.append(_private_text_candidate(
                    source_ref=source_ref,
                    locator=f"SHEET_{sheet_index}:{cell.coordinate}",
                    family=family,
                    raw_text=str(_json_value(cell.value)),
                    method="XLSX_HEADER_BOUND_CELL",
                    formula_text=cell.value if isinstance(cell.value, str) and cell.value.startswith("=") else None,
                    cached_display_value=display_value,
                    candidate_role=_candidate_role(family, header_text, workbook=True),
                    identity_component=component,
                ) | {"source_header_raw_text": header_text})

        if sheet_formula_count:
            template_class = "XLSX_FORMULA_SUMMARY"
        elif nonempty_count:
            template_class = "XLSX_MANUAL_DETAIL"
        elif media["media_count"]:
            template_class = "XLSX_IMAGE_EMBEDDED"
        else:
            template_class = "UNRECOGNIZED_XLSX_TEMPLATE"
        sheet_records.append({
            "sheet_index": sheet_index,
            "private_sheet_name": sheet.title,
            "nonempty_cell_count": nonempty_count,
            "formula_cell_count": sheet_formula_count,
            "cached_formula_display_count": sheet_cached_count,
            "manual_numeric_cell_count": sheet_manual_numeric,
            "template_class": template_class,
            "parser_strategy": TEMPLATE_STRATEGIES.get(template_class, "QUARANTINE_ONLY"),
            "human_confirmation_required": True,
        })
        formula_count += sheet_formula_count
        cached_formula_count += sheet_cached_count
        manual_numeric_count += sheet_manual_numeric

    template_classes = sorted({item["template_class"] for item in sheet_records})
    quarantined = [
        {"locator": f"SHEET_{item['sheet_index']}", "reason": "UNRECOGNIZED_TEMPLATE"}
        for item in sheet_records if item["template_class"] not in TEMPLATE_STRATEGIES
    ]
    return {
        "source_ref": source_ref,
        "format": "XLSX",
        "sheet_count": len(sheet_records),
        "sheets": sheet_records,
        "field_candidates": candidates,
        "template_classes": template_classes,
        "formula_cell_count": formula_count,
        "cached_formula_display_count": cached_formula_count,
        "manual_numeric_cell_count": manual_numeric_count,
        "formula_and_display_separated": formula_count > 0 and formula_count == cached_formula_count,
        **media,
        "quarantined_components": quarantined,
        "ocr_performed": False,
        "ocr_final_fact_count": 0,
    }


def _secure_write_json(path: Path, payload: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True, mode=0o700)
    path.parent.chmod(0o700)
    temporary = path.with_suffix(path.suffix + ".tmp")
    temporary.write_text(json.dumps(payload, ensure_ascii=False, indent=2, sort_keys=True) + "\n", encoding="utf-8")
    temporary.chmod(0o600)
    temporary.replace(path)
    path.chmod(0o600)


def run_private_scan(
    *, raw_root: Path = RAW_INBOX, output_path: Path = PRIVATE_MANIFEST_PATH,
) -> dict[str, Any]:
    raw_before = _stat_snapshot(raw_root) if raw_root.is_dir() else {}
    package, selector_status, match_count = resolve_authority_package(raw_root)
    if package is None:
        raise RegistrationError(f"authority package selection failed: {selector_status} ({match_count})")
    package_before = _stat_snapshot(package)
    package_hash_before = _sha256_file(package)
    source_records: list[dict[str, Any]] = []
    with zipfile.ZipFile(package) as archive:
        members = _business_members(archive)
        pdf_count = sum(Path(item.filename).suffix.lower() == ".pdf" for item in members)
        workbook_count = sum(Path(item.filename).suffix.lower() in {".xlsx", ".xlsm"} for item in members)
        if (len(members), pdf_count, workbook_count) != (
            EXPECTED_SOURCE_COUNT, EXPECTED_PDF_COUNT, EXPECTED_WORKBOOK_COUNT,
        ):
            raise RegistrationError("authority package no longer has the required 8 PDF plus 1 workbook shape")
        for index, member in enumerate(members, start=1):
            source_ref = f"S06P1-SRC-{index:03d}"
            payload = archive.read(member)
            suffix = Path(member.filename).suffix.lower()
            if suffix == ".pdf":
                inspection = inspect_pdf_payload(payload, source_ref)
                role = "AUTHORITATIVE_PROJECT_COST_PDF"
            elif suffix in {".xlsx", ".xlsm"}:
                inspection = inspect_workbook_payload(payload, source_ref)
                role = "AUTHORITATIVE_PROJECT_COST_WORKBOOK"
            else:  # pragma: no cover - shape gate rejects this
                raise RegistrationError(f"unsupported source format for {source_ref}")
            source_records.append({
                "source_ref": source_ref,
                "archive_order": index,
                "private_member_name": member.filename,
                "private_member_name_sha256": _sha256_text(member.filename),
                "private_member_sha256": _sha256_bytes(payload),
                "private_member_size_bytes": len(payload),
                "source_role": role,
                "integrity_status": "READABLE_HASHED",
                "inspection": inspection,
            })
    package_hash_after = _sha256_file(package)
    package_after = _stat_snapshot(package)
    raw_after = _stat_snapshot(raw_root)
    payload = {
        "schema_version": PRIVATE_SCHEMA_VERSION,
        "classification": "PRIVATE_RAW_DERIVED_DO_NOT_COMMIT",
        "project_id": "KMFA",
        "target_release": "v1.5",
        "phase_id": RUN_PHASE_ID,
        "roadmap_phase_id": ROADMAP_PHASE_ID,
        "task_id": TASK_ID,
        "acceptance_id": ACCEPTANCE_ID,
        "generated_at": datetime.now().astimezone().isoformat(timespec="seconds"),
        "private_raw_root": str(raw_root),
        "private_package_path": str(package),
        "private_package_name": package.name,
        "private_package_sha256": package_hash_before,
        "selector_status": selector_status,
        "selector_match_count": match_count,
        "source_count": len(source_records),
        "pdf_count": sum(item["inspection"]["format"] == "PDF" for item in source_records),
        "workbook_count": sum(item["inspection"]["format"] == "XLSX" for item in source_records),
        "source_records": source_records,
        "raw_root_before": raw_before,
        "raw_root_after": raw_after,
        "package_before": package_before,
        "package_after": package_after,
        "package_hash_before": package_hash_before,
        "package_hash_after": package_hash_after,
        "raw_root_stat_unchanged": raw_before == raw_after,
        "package_stat_unchanged": package_before == package_after,
        "package_hash_unchanged": package_hash_before == package_hash_after,
        "raw_read_performed": True,
        "raw_list_performed": True,
        "raw_stat_performed": True,
        "raw_hash_performed": True,
        "raw_parse_performed": True,
        "raw_write_performed": False,
        "raw_delete_performed": False,
        "raw_move_performed": False,
        "raw_rename_performed": False,
        "raw_overwrite_performed": False,
        "raw_mutation_performed": False,
        "ocr_final_fact_count": 0,
        "golden_value_confirmed_count": 0,
        "s06_p2_started": False,
        "formal_report_generated": False,
        "github_upload_performed": False,
        "app_reinstall_performed": False,
        "business_execution_performed": False,
    }
    validate_private_payload(payload)
    _secure_write_json(output_path, payload)
    return payload


def read_private_payload(path: Path = PRIVATE_MANIFEST_PATH) -> dict[str, Any]:
    value = json.loads(path.read_text(encoding="utf-8"))
    if not isinstance(value, dict):
        raise RegistrationError("private registration manifest must be a JSON object")
    validate_private_payload(value)
    return value


def validate_private_payload(payload: dict[str, Any]) -> None:
    if payload.get("schema_version") != PRIVATE_SCHEMA_VERSION:
        raise RegistrationError("private schema mismatch")
    if (payload.get("source_count"), payload.get("pdf_count"), payload.get("workbook_count")) != (9, 8, 1):
        raise RegistrationError("private source shape mismatch")
    records = payload.get("source_records")
    if not isinstance(records, list) or len(records) != 9:
        raise RegistrationError("nine private source records are required")
    refs = [item.get("source_ref") for item in records]
    if len(set(refs)) != 9 or not all(re.fullmatch(r"S06P1-SRC-\d{3}", str(ref)) for ref in refs):
        raise RegistrationError("private source refs are invalid")
    if not all(item.get("integrity_status") == "READABLE_HASHED" for item in records):
        raise RegistrationError("every private source must be readable and hashed")
    for item in records:
        inspection = item.get("inspection")
        candidates = inspection.get("field_candidates") if isinstance(inspection, dict) else None
        if not isinstance(candidates, list) or not candidates:
            raise RegistrationError("every source needs private field candidates")
        locator_families: dict[str, set[str]] = {}
        for candidate in candidates:
            if not isinstance(candidate, dict):
                raise RegistrationError("private candidate must be an object")
            family = candidate.get("field_family")
            locator = candidate.get("source_locator")
            if family not in FIELD_FAMILIES:
                raise RegistrationError("private candidate family is invalid")
            if not isinstance(locator, str) or not locator:
                raise RegistrationError("private candidate locator is missing")
            if candidate.get("candidate_role") not in _CANDIDATE_ROLES:
                raise RegistrationError("private candidate role is invalid")
            locator_families.setdefault(locator, set()).add(family)
            if inspection.get("format") == "PDF" and not locator.startswith("PAGE_1:"):
                raise RegistrationError("supporting PDF pages cannot become golden candidates")
            header = _normalize(candidate.get("source_header_raw_text"))
            if family == "GROSS_PROFIT" and header == _normalize("毛利率"):
                raise RegistrationError("margin header cannot create gross-profit candidate")
            if _normalize(candidate.get("raw_text")) in _WORKBOOK_SUMMARY_LABELS:
                raise RegistrationError("workbook summary label cannot become a field candidate")
        if any({"CONTRACT_AMOUNT", "TOTAL_EXPENDITURE"} <= families for families in locator_families.values()):
            raise RegistrationError("one source locator cannot be both contract amount and total expenditure")
    if not all(payload.get(key) is True for key in (
        "raw_root_stat_unchanged", "package_stat_unchanged", "package_hash_unchanged",
    )):
        raise RegistrationError("raw stat or content changed during scan")
    if any(payload.get(key) is not False for key in (
        "raw_write_performed", "raw_delete_performed", "raw_move_performed",
        "raw_rename_performed", "raw_overwrite_performed", "raw_mutation_performed",
        "s06_p2_started", "formal_report_generated", "github_upload_performed",
        "app_reinstall_performed", "business_execution_performed",
    )):
        raise RegistrationError("out-of-scope mutation or downstream action detected")
    if payload.get("ocr_final_fact_count") != 0 or payload.get("golden_value_confirmed_count") != 0:
        raise RegistrationError("S06-P1 cannot create OCR final facts or confirmed golden values")


def public_projection(payload: dict[str, Any]) -> dict[str, dict[str, Any]]:
    validate_private_payload(payload)
    source_rows: list[dict[str, Any]] = []
    family_counts: Counter[str] = Counter()
    family_sources: dict[str, set[str]] = {name: set() for name in FIELD_FAMILIES}
    template_counts: Counter[str] = Counter()
    quarantined_component_count = 0
    formula_count = cached_formula_count = workbook_media_count = 0
    textless_page_count = 0
    candidate_roles: Counter[str] = Counter()
    contract_total_locator_collision_count = 0
    supporting_pdf_promoted_candidate_count = 0
    margin_header_gross_profit_candidate_count = 0
    workbook_summary_candidate_count = 0
    for item in payload["source_records"]:
        inspection = item["inspection"]
        candidates = inspection.get("field_candidates", [])
        locator_families: dict[str, set[str]] = {}
        for candidate in candidates:
            family = candidate["field_family"]
            if family in family_sources:
                family_counts[family] += 1
                family_sources[family].add(item["source_ref"])
            candidate_roles[candidate["candidate_role"]] += 1
            locator_families.setdefault(candidate["source_locator"], set()).add(family)
            if inspection["format"] == "PDF" and not candidate["source_locator"].startswith("PAGE_1:"):
                supporting_pdf_promoted_candidate_count += 1
            if family == "GROSS_PROFIT" and _normalize(candidate.get("source_header_raw_text")) == _normalize("毛利率"):
                margin_header_gross_profit_candidate_count += 1
            if _normalize(candidate.get("raw_text")) in _WORKBOOK_SUMMARY_LABELS:
                workbook_summary_candidate_count += 1
        contract_total_locator_collision_count += sum(
            {"CONTRACT_AMOUNT", "TOTAL_EXPENDITURE"} <= families
            for families in locator_families.values()
        )
        classes = [inspection["template_class"]] if inspection["format"] == "PDF" else inspection["template_classes"]
        template_counts.update(classes)
        quarantined_component_count += len(inspection.get("quarantined_components", []))
        textless_page_count += int(inspection.get("textless_page_count", 0))
        formula_count += int(inspection.get("formula_cell_count", 0))
        cached_formula_count += int(inspection.get("cached_formula_display_count", 0))
        workbook_media_count += int(inspection.get("media_count", 0))
        source_rows.append({
            "source_ref": item["source_ref"],
            "format": inspection["format"],
            "source_role": item["source_role"],
            "integrity_status": item["integrity_status"],
            "private_locator_recorded": True,
            "private_hash_recorded": True,
            "private_raw_text_recorded": bool(candidates),
            "field_candidate_count": len(candidates),
            "template_classes": classes,
            "template_identified": all(value in TEMPLATE_STRATEGIES for value in classes),
            "human_confirmation_required": True,
            "golden_value_confirmed": False,
            "raw_name_committed": False,
            "raw_hash_committed": False,
            "raw_text_committed": False,
            "raw_value_committed": False,
            "sheet_name_committed": False,
            "formula_text_committed": False,
        })
    coverage_rows = [
        {
            "field_family": family,
            "candidate_count": family_counts[family],
            "source_coverage_count": len(family_sources[family]),
            "private_original_text_and_locator_recorded": family_counts[family] > 0,
            "candidate_status": "CANDIDATE_NOT_FINAL",
            "human_confirmation_required": True,
        }
        for family in FIELD_FAMILIES
    ]
    registration = {
        "schema_version": PUBLIC_SCHEMA_VERSION,
        "project_id": "KMFA",
        "target_release": "v1.5",
        "stage_id": "S06",
        "phase_id": RUN_PHASE_ID,
        "roadmap_phase_id": ROADMAP_PHASE_ID,
        "source_count": len(source_rows),
        "pdf_count": sum(row["format"] == "PDF" for row in source_rows),
        "workbook_count": sum(row["format"] == "XLSX" for row in source_rows),
        "source_records": source_rows,
        "all_sources_readable_hashed_private_only": all(row["integrity_status"] == "READABLE_HASHED" for row in source_rows),
        "raw_root_stat_unchanged": payload["raw_root_stat_unchanged"],
        "package_stat_unchanged": payload["package_stat_unchanged"],
        "package_hash_unchanged": payload["package_hash_unchanged"],
        "private_evidence_token": "PRIVATE_RUNTIME_TOKEN::V015_S06_P1_AUTHORITY_REGISTRATION",
        "public_raw_name_count": 0,
        "public_raw_hash_count": 0,
        "public_raw_text_count": 0,
        "public_raw_value_count": 0,
        "public_sheet_name_count": 0,
    }
    coverage = {
        "schema_version": "kmfa.v015.s06p1.field_candidate_coverage_public_safe.v1",
        "field_family_count": len(coverage_rows),
        "covered_field_family_count": sum(row["candidate_count"] > 0 for row in coverage_rows),
        "candidate_count": sum(row["candidate_count"] for row in coverage_rows),
        "coverage": coverage_rows,
        "all_candidates_private_only": True,
        "original_display_text_locator_private_only": True,
        "golden_value_confirmed_count": 0,
        "ocr_final_fact_count": 0,
        "candidate_role_counts": dict(sorted(candidate_roles.items())),
        "contract_total_locator_collision_count": contract_total_locator_collision_count,
        "supporting_pdf_promoted_candidate_count": supporting_pdf_promoted_candidate_count,
        "margin_header_gross_profit_candidate_count": margin_header_gross_profit_candidate_count,
        "workbook_summary_candidate_count": workbook_summary_candidate_count,
        "candidate_semantic_quality_passed": all(value == 0 for value in (
            contract_total_locator_collision_count,
            supporting_pdf_promoted_candidate_count,
            margin_header_gross_profit_candidate_count,
            workbook_summary_candidate_count,
        )),
    }
    template = {
        "schema_version": "kmfa.v015.s06p1.template_difference_public_safe.v1",
        "observed_template_class_count": len(template_counts),
        "observed_template_classes": [
            {
                "template_class": name,
                "source_or_sheet_count": template_counts[name],
                "parser_strategy": TEMPLATE_STRATEGIES.get(name, "QUARANTINE_ONLY"),
                "human_confirmation_required": True,
            }
            for name in sorted(template_counts)
        ],
        "all_observed_template_classes_have_strategy": all(name in TEMPLATE_STRATEGIES for name in template_counts),
        "quarantined_component_count": quarantined_component_count,
        "textless_page_count": textless_page_count,
        "unknown_template_source_count": sum(not row["template_identified"] for row in source_rows),
        "formula_cell_count": formula_count,
        "cached_formula_display_count": cached_formula_count,
        "formula_and_display_values_separated": formula_count > 0 and formula_count == cached_formula_count,
        "workbook_embedded_media_count": workbook_media_count,
        "ocr_performed": False,
        "ocr_final_fact_count": 0,
        "unknown_template_auto_parse_allowed": False,
    }
    return {"registration": registration, "coverage": coverage, "template": template}


def main(argv: Iterable[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="KMFA v1.5 S06-P1 private authority source scanner")
    parser.add_argument("--private-scan", action="store_true")
    parser.add_argument("--private-output", type=Path, default=PRIVATE_MANIFEST_PATH)
    args = parser.parse_args(list(argv) if argv is not None else None)
    if not args.private_scan:
        parser.error("--private-scan is required; public artifacts are produced by the builder")
    result = run_private_scan(output_path=args.private_output)
    projection = public_projection(result)
    print(json.dumps({
        "status": "PASS",
        "source_count": projection["registration"]["source_count"],
        "pdf_count": projection["registration"]["pdf_count"],
        "workbook_count": projection["registration"]["workbook_count"],
        "field_family_count": projection["coverage"]["field_family_count"],
        "covered_field_family_count": projection["coverage"]["covered_field_family_count"],
        "template_class_count": projection["template"]["observed_template_class_count"],
        "raw_mutation_performed": False,
    }, sort_keys=True))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
